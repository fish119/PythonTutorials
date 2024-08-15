from datetime import datetime
import os
import openpyxl

# 获取当前文件夹中的所有Excel文件
excel_files = [file for file in os.listdir() if file.endswith(".xlsx")]

for file in excel_files:
    wb = openpyxl.load_workbook(file)
    if file != "01.xlsx":

        # 删除Sheet2和Sheet3
        for sheet_name in ["Sheet2", "Sheet3"]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

    # 将第一个工作表的名称更改为Sheet1（如果不是Sheet1）
    first_sheet = wb.sheetnames[0]
    if first_sheet != "Sheet1":
        ws = wb[first_sheet]
        ws.title = "Sheet1"

        # 遍历每个工作表
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 遍历每行和每列
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                # 如果单元格内容是字符串类型，去除多余空格
                if isinstance(cell.value, str):
                    cell.value = cell.value.strip()

                if isinstance(cell.value, datetime) and cell.value.year == 2002:
                    # 替换日期中的年份
                    new_date = cell.value.replace(year=2024)
                    cell.value = new_date

    # 保存更改后的Excel文件
    wb.save(file)

    print(f"处理完成：{file}")
