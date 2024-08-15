import os
from openpyxl import load_workbook


def calculate_and_copy_values(input_file):
    # 使用data_only=True加载工作簿以计算公式
    wb_calculated = load_workbook(input_file, data_only=True)

    # 再次加载工作簿以保留原始公式
    wb = load_workbook(input_file)

    # 获取Sheet2和Sheet3
    if "Sheet2" in wb_calculated.sheetnames:
        sheet2_calculated = wb_calculated["Sheet2"]

        # 如果Sheet3存在，先删除它
        if "Sheet3" in wb.sheetnames:
            wb.remove(wb["Sheet3"])

        # 创建Sheet3
        wb.create_sheet("Sheet3")
        sheet3 = wb["Sheet3"]

        # 复制Sheet2的数据到Sheet3
        for row in sheet2_calculated.iter_rows(values_only=True):
            sheet3.append(row)

        # 保存修改后的Excel文件
        wb.save(input_file)

    # 关闭工作簿
    wb_calculated.close()


# # 处理当前文件夹下的04.xlsx
# input_file = "07.xlsx"
# print(f"Processing {input_file}...")
# calculate_and_copy_values(input_file)

excel_files = [file for file in os.listdir() if file.endswith(".xlsx")]

for file in excel_files:
    if file != "01.xlsx":  # 排除01.xlsx文件本身
        print(f"Processing {file}...")
        calculate_and_copy_values(file)