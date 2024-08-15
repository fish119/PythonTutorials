from openpyxl import load_workbook


def copy_values_to_another_sheet(input_file):
    # 加载Excel文件
    wb = load_workbook(input_file)

    # 创建一个临时工作簿，用于计算Sheet2中的值
    temp_wb = load_workbook(input_file, data_only=True)

    # 获取Sheet2和Sheet3
    if "Sheet2" in wb.sheetnames:
        sheet2 = wb["Sheet2"]
        sheet2_data_only = temp_wb["Sheet2"]

        # 如果Sheet3存在，先删除它
        if "Sheet3" in wb.sheetnames:
            wb.remove(wb["Sheet3"])

        # 创建Sheet3
        wb.create_sheet("Sheet3")
        sheet3 = wb["Sheet3"]

        # 复制Sheet2的数据到Sheet3
        for row in sheet2_data_only.iter_rows(values_only=True):
            sheet3.append(row)

        # 关闭临时工作簿
        temp_wb.close()

        # 保存修改后的Excel文件
        wb.save(input_file)


# 处理当前文件夹下的01.xlsx
input_file = "01.xlsx"
print(f"Processing {input_file}...")
copy_values_to_another_sheet(input_file)
