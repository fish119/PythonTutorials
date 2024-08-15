import os
import openpyxl
import re

# 获取当前文件夹中的所有Excel文件
excel_files = [file for file in os.listdir() if file.endswith(".xlsx")]

# 读取01.xlsx中的Sheet3
wb_source = openpyxl.load_workbook("01.xlsx")

# 复制Sheet3到其他Excel文件中
for file in excel_files:
    if file != "01.xlsx":  # 排除01.xlsx文件本身
        wb_dest = openpyxl.load_workbook(file)
        for sheet_name in ["Sheet2"]:
            if sheet_name in wb_source.sheetnames:
                ws_source = wb_source[sheet_name]
                ws_dest = wb_dest.create_sheet(title=sheet_name)

                # 复制源工作表数据
                for row in ws_source.iter_rows(values_only=True):
                    ws_dest.append(row)

        # 保存更改后的Excel文件
        wb_dest.save(file)

        print(f"复制完成：{file}")

# 替换所有Excel文件中的公式中的'[01.xlsx]'为''
for file in excel_files:
    wb = openpyxl.load_workbook(file)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.data_type == "f" and "[01.xlsx]" in str(cell.value):
                    cell.value = re.sub(r"\[01.xlsx\]", "", str(cell.value))

    # 保存更改后的Excel文件
    wb.save(file)

    print(f"替换完成：{file}")
